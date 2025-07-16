from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import load_workbook
import pandas as pd
import shutil
from Transform_journey_results import read_csv_files
from openpyxl.utils import get_column_letter, column_index_from_string
PATH = 'C:/Users/choc-/OneDrive/Escritorio/WA_Campaigns_Report_Creation/journey_files/inputs'
DB_PATH = 'C:/Users/choc-/OneDrive/Escritorio/WA_Campaigns_Report_Creation/data/db'

import shutil


# def read_campaing_data()
    


def find_conversations(date: str) -> pd.DataFrame:
    
    conversations_df = pd.read_csv(os.path.join(DB_PATH, f'ARG_Historic_Conversations.csv'))

    conversations_df['conversation_date'] = pd.to_datetime(conversations_df['conversation_date'])
    start_date = pd.to_datetime(date)
    end_date = start_date + pd.Timedelta(hours=24)
    filtered_df = conversations_df[
        (conversations_df['conversation_date'] >= start_date) &
        (conversations_df['conversation_date'] < end_date)
    ]
    return filtered_df


def transform_data():
    # Leer CSV
    journey_results_df = read_csv_files(
        'C:/Users/choc-/OneDrive/Escritorio/WA_Campaigns_Report_Creation/jouneys_files/inputs/1232_Resultados.csv'
    )

    if 'shot' in journey_results_df.columns:
        journey_results_df.loc[:, 'shot'] = (
            journey_results_df['shot']
            .str.lower()
            .str.replace(' ', '', regex=False)
            .str.replace('_', '', regex=False)
        )

    if journey_results_df.empty:
        print('No hay datos para escribir en Excel.')
        return

    # Copiar archivo
    original_excel_path = 'C:/Users/choc-/OneDrive/Escritorio/WA_Campaigns_Report_Creation/src/utils/Template_WA.xlsx'
    new_excel_path = 'C:/Users/choc-/OneDrive/Escritorio/WA_Campaigns_Report_Creation/jouneys_files/outputs/WA_Campaigns_Report_Creation.xlsx'
    shutil.copy(original_excel_path, new_excel_path)

    wb = load_workbook(new_excel_path)
    ws = wb['data']

    table_name = 'WA_journey'
    if table_name not in ws.tables:
        print(f"No se encontró la tabla llamada {table_name}")
        return
    table = ws.tables[table_name]

    ref = table.ref
    start_cell, end_cell = ref.split(':')
    min_col_letter = ''.join(filter(str.isalpha, start_cell))
    min_row = int(''.join(filter(str.isdigit, start_cell)))
    max_col_letter = ''.join(filter(str.isalpha, end_cell))
    max_row = int(''.join(filter(str.isdigit, end_cell)))

    # Encontrar última fila con datos dentro de la tabla para evitar huecos
    # Itera desde la fila siguiente al encabezado hasta max_row para buscar fila vacía
    last_data_row = min_row  # empieza asumiendo sólo encabezado

    for row_idx in range(min_row + 1, max_row + 1):
        # Revisar si la fila tiene datos en la primera columna de la tabla
        cell_val = ws[f"{min_col_letter}{row_idx}"].value
        if cell_val is not None and str(cell_val).strip() != "":
            last_data_row = row_idx

    start_row = last_data_row + 1
    start_col = column_index_from_string(min_col_letter)

    # Escribir las filas justo después de la última fila con datos para evitar fila en blanco
    for r_idx, row in enumerate(dataframe_to_rows(journey_results_df, index=False, header=False), start=start_row):
        for c_idx, value in enumerate(row, start=start_col):
            ws.cell(row=r_idx, column=c_idx, value=value)

    # Actualizar rango de tabla (extiende el max_row)
    new_max_row = max(max_row, start_row + len(journey_results_df) - 1)
    table.ref = f"{min_col_letter}{min_row}:{max_col_letter}{new_max_row}"

    wb.save(new_excel_path)
    print(f"Copia guardada en: {new_excel_path} con filas agregadas correctamente sin espacios.")

transform_data()

